#!/usr/bin/env python3
import json
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader


def normalize_text(text: str) -> str:
    lines = [line.strip() for line in text.replace("\u3000", " ").splitlines()]
    lines = [line for line in lines if line]
    return "\n".join(lines).strip()


def extract_xlsx(file_path: Path) -> dict:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    chunks: list[str] = []
    row_count = 0

    for sheet in workbook.worksheets:
        chunks.append(f"[Sheet] {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = [str(cell).strip() for cell in row if cell not in (None, "")]
            if not values:
                continue
            row_count += 1
            chunks.append(" | ".join(values))

    text = normalize_text("\n".join(chunks))
    return {
        "kind": "xlsx",
        "page_count": None,
        "row_count": row_count,
        "text_source": "xlsx",
        "text": text,
    }


def extract_pdf_with_pypdf(file_path: Path) -> tuple[int, list[str]]:
    reader = PdfReader(str(file_path))
    pages = []
    for page in reader.pages:
        pages.append(normalize_text(page.extract_text() or ""))
    return len(reader.pages), pages


def extract_pdf_with_vision(file_path: Path) -> tuple[int, list[str]]:
    script = Path(__file__).with_name("pdf_vision_ocr.swift")
    result = subprocess.run(
        ["swift", str(script), str(file_path)],
        check=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
    )
    payload = json.loads(result.stdout)
    return payload["pageCount"], [normalize_text(page["text"]) for page in payload["pages"]]


def extract_pdf(file_path: Path) -> dict:
    page_count, pypdf_pages = extract_pdf_with_pypdf(file_path)
    pypdf_text = normalize_text("\n\n".join(pypdf_pages))

    if len(pypdf_text.replace("\n", "")) >= 80:
        return {
            "kind": "pdf",
            "page_count": page_count,
            "text_source": "pypdf",
            "text": pypdf_text,
        }

    ocr_page_count, ocr_pages = extract_pdf_with_vision(file_path)
    ocr_text = normalize_text("\n\n".join(ocr_pages))

    if len(ocr_text.replace("\n", "")) < 80:
        raise RuntimeError(f"failed to extract usable text from pdf: {file_path}")

    return {
        "kind": "pdf",
        "page_count": ocr_page_count,
        "text_source": "vision_ocr",
        "text": ocr_text,
    }


def main() -> int:
    if len(sys.argv) != 3:
        print("usage: extract_case_document.py <input_file> <output_json>", file=sys.stderr)
        return 1

    input_file = Path(sys.argv[1]).resolve()
    output_file = Path(sys.argv[2]).resolve()

    if not input_file.exists():
        print(f"input file not found: {input_file}", file=sys.stderr)
        return 1

    suffix = input_file.suffix.lower()
    if suffix == ".xlsx":
        payload = extract_xlsx(input_file)
    elif suffix == ".pdf":
        payload = extract_pdf(input_file)
    else:
        print(f"unsupported file type: {input_file}", file=sys.stderr)
        return 1

    payload["file_name"] = input_file.name
    payload["absolute_path"] = str(input_file)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    temp_file = output_file.with_suffix(output_file.suffix + ".tmp")
    temp_file.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    temp_file.replace(output_file)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
